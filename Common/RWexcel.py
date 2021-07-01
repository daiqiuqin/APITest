# #-*-coding:utf-8 -*-
# #@Time :2021/1/21 14:52
# #@Author :daiqiuqin
# #@File  :RWexcel.py

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.styles import Font,Color,Alignment,PatternFill,Border, Side
from Common.handle_path import datas_dir,reports_dir
import os

class ReadExcel:
    def __init__(self,readFile_path,readSheet_name):
        self.file_path=readFile_path
        self.sheet_name=readSheet_name
        self.wb=load_workbook(self.file_path)
        # sheet.rows为生成器, 里面是每一行的数据，每一行又由一个tuple包裹。
        # sheet.columns类似，不过里面是每个tuple是每一列的单元格。
        # 因为按行，所以返回A1, B1, C1这样的顺序
        self.sheet=self.wb[self.sheet_name]

    def read_titles(self):
        '''读取sheet页第一行作为titles
        :return:列表表头
        :rtype:List
        '''
        titles = []
        for colum in self.sheet.columns:
            titles.append(colum[0].value)
        return titles

    def read_all_datas(self):
        '''读取sheet测试用例
        :return:以字典组成的列表形式，返回每行的测试用例,除去表头
        '''
        all_datas = []
        titles = self.read_titles()
        for item in list(self.sheet.rows)[1:]:
            values = []
            for val in item:
                values.append(val.value)
            res = dict(zip(titles, values))
            all_datas.append(res)
        self.wb.close()
        return all_datas

    def get_max_row(self):
        return self.sheet.max_row

    def get_max_column(self):
        return self.sheet.max_column



class WriteExcel():
    def __init__(self,writeFile_path,writeSheet_name):
        self.file_path=writeFile_path
        self.sheet_name=writeSheet_name
        self.wb = load_workbook(self.file_path)
        self.sheet = self.wb[self.sheet_name]

    @staticmethod
    def __set_font(bold=False,color='000000'):
        '''设置字体样式,默认：“宋体”，11，不加粗，无斜体，无删除线，黑色。可修改是否加粗,字体颜色
                :return: 字体、对齐方式、填充样式
                '''
        # 设置字体为“宋体”，大小为11，bold为加粗，italic为斜体，strike为删除线，颜色为黑色
        fontValue = Font(u'宋体', size=11, bold=bold, italic=False, strike=False, color=color)
        return fontValue

    @staticmethod
    def __set_fill(start_color='FFFFFF'):
        '''设置填充样式,默认前景色为白，后景色为白。可修改前景色。
                :return: 字体、对齐方式、填充样式
                '''
        # fill_type:solid为纯色填充，start_color代表前景色，end_color是背景色
        fillValue = PatternFill(fill_type='solid', start_color=start_color, end_color='FFFFFF')
        return fillValue

    @staticmethod
    def __set_align():
        '''设置对齐样式,默认左对齐，居中。
                :return: 字体、对齐方式、填充样式
                '''
        # horizontal代表水平方向，vertical代表垂直方向，wrap_text代码自动换行
        alignValue = Alignment(horizontal='left', vertical='center', wrap_text=True)
        return alignValue

    @staticmethod
    def __set_border():
        '''设置边框样式,默认：黑色细边框
                        :return: 字体、对齐方式、填充样式
                        '''
        # 设置边框样式
        borderValue = Border(left=Side(border_style='thin', color='000000'),
                             right=Side(border_style='thin', color='000000'),
                             top=Side(border_style='thin', color='000000'),
                             bottom=Side(border_style='thin', color='000000'))
        return borderValue

    def excelWrite(self,row,id,title,method,request_url,request_body,expected,realResult,testResult,output_file):
        '''用于测试结果的写入及excel格式的设置
        output_file:输出excel的地址
        row:读取第几行
        id:用例id
        title:用例title
        method:请求方法
        request_url:请求url
        request_body:请求内容
        expected：预期结果
        response_text：响应结果
        testResult：执行结果，Pass or Fail
        '''
        # 创建workbook对象
        wb = load_workbook(output_file)
        wb.active
        sh = wb[self.sheet_name]

        datas={}
        datas['id']=id
        datas['title'] = title
        datas['method']=method
        datas['request_url'] = request_url
        datas['request_body'] = request_body
        datas['expected'] = expected
        datas['realResult']=realResult
        testResult=testResult
        # print(datas)

        # 用例数据回写
        col = 1
        titles = ['id', 'title','method','request_url', 'request_body','expected', 'realResult', 'testResult']
        length=len(titles)
        # print('length:'+str(length))
        # print(titles)
        for item in titles:

            # 设置全局的对齐方式、边框样式
            alignValue = self.__set_align()
            borderValue = self.__set_border()

            # 写表头
            # print("item:"+item)
            sh.cell(1, col).value = item

            # 设置表头字体:加粗、白色
            fontValue=self.__set_font('True','FFFFFF')
            # 设置表头前景色为蓝色
            fillValue=self.__set_fill('58ACFA')
            # 设置表头格式
            sh.cell(1, col).font=fontValue
            sh.cell(1, col).alignment=alignValue
            sh.cell(1, col).fill = fillValue
            sh.cell(1, col).border = borderValue
            # 调整表头行高
            sh.row_dimensions[1].height = 20

            # 用例回写
            if col<=length-1:
                sh.cell(row + 1, col).value = datas[item]
                # print(datas[item])
            elif col==length:
                sh.cell(row + 1, length).value = testResult
                if testResult.upper()=='FAIL' or testResult=='不通过':
                    fillValue1=self.__set_fill('FF0000')
                    sh.cell(row + 1, length).fill=fillValue1
                elif testResult.upper()=='PASS' or testResult=='通过':
                    fillValue2=self.__set_fill('3ADF00')
                    sh.cell(row + 1, length).fill=fillValue2

            # 设置内容格式
            fontValue1= self.__set_font()
            sh.cell(row + 1, col).font=fontValue1
            sh.cell(row + 1, col).alignment=alignValue
            sh.cell(row + 1, col).border=borderValue
            col += 1
            # 调整内容行高
            sh.row_dimensions[row+1].height = 120

            # 调整各个字段的列宽,根据表头来设定列宽，可自己调整
            lableList = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']
            try:
                sh.column_dimensions[lableList[titles.index("id")]].width = 15.0
                sh.column_dimensions[lableList[titles.index("title")]].width = 20.0
                sh.column_dimensions[lableList[titles.index("method")]].width = 8.0
                sh.column_dimensions[lableList[titles.index("request_url")]].width = 50.0
                sh.column_dimensions[lableList[titles.index("request_body")]].width = 50.0
                sh.column_dimensions[lableList[titles.index("expected")]].width = 20.0
                sh.column_dimensions[lableList[titles.index("realResult")]].width = 20.0
                sh.column_dimensions[lableList[titles.index("testResult")]].width = 12.0
            except Exception as e:
                print("测试用例列表字段不符合规范")
                print(e)
        wb.save(output_file)
        wb.close()

def Wexcel(output_file,readSheet,row, id, title, method,request_url, request_body, expected, response_text, testResult):
    '''用于excel写入的前置处理：1、判断表格是否存在，如果不存在，则新建 2、判断sheet是否存在，不存在则新建
    output_file:输出excel的地址
    readsheet:读取的sheet名称
    row:读取第几行
    id:用例id
    title:用例title
    method:请求方法
    request_url:请求url
    request_body:请求内容
    expected：预期结果
    response_text：响应结果
    testResult：执行结果，Pass or Fail
    '''
    # 判断是否存在表格；不存在则创建新表格，存在则直接使用表格
    if not os.path.exists(output_file):
        # 实例化对象
        wb = Workbook()
        # 新建一个xlsx，要激活工作簿
        wb.active
        # 在索引-1位置增加sheet，-1表示倒数第二
        sh = wb.create_sheet(readSheet,-1)
        wb.save(output_file)
        wb.close()
    else:
        wb = load_workbook(output_file)
        wb.active
        # 判断sheet是否存在，不存在则新建一个
        if readSheet not in wb.sheetnames:
            sh = wb.create_sheet(readSheet,-1)
            print("{}不存在，新建一个".format(readSheet))
            wb.save(output_file)
            wb.close()
    try:
        cc = WriteExcel(output_file, readSheet)
        cc.excelWrite(row, id, title, method,request_url, request_body, expected, response_text, testResult, output_file)
    except IOError as e:
        print(e)

if __name__=='__main__':
    # # 读取测试用例
    readExcel=os.path.join(datas_dir,'智慧服务接口测试.xlsx')
    readSheet = '查询病历'
    excel = ReadExcel(readExcel, readSheet)
    # titles=excel.read_titles()
    alldatas=excel.read_all_datas()
    print(alldatas)
    # mr = excel.get_max_row()
    # # mc = excel.get_max_column()
    # print('用例行数：' + str(mr))


